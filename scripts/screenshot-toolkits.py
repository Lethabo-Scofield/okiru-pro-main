#!/usr/bin/env python3
"""
Automated BBBEE toolkit screenshot pipeline.

Renders every sheet of every toolkit xlsx to a PNG using openpyxl + Pillow
and produces docs/toolkits/screenshots/index.md.

Run:  py scripts/screenshot-toolkits.py
"""

import os
import re
import sys
import time
import textwrap
from pathlib import Path

try:
    import openpyxl
    from openpyxl.utils import get_column_letter, column_index_from_string
    from openpyxl.styles.fills import PatternFill
except ImportError:
    sys.exit("openpyxl not found — run: py -m pip install openpyxl")

try:
    from PIL import Image, ImageDraw, ImageFont
except ImportError:
    sys.exit("Pillow not found — run: py -m pip install Pillow")

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

TOOLKITS_DIR = Path("docs/toolkits")
OUT_DIR       = Path("docs/toolkits/screenshots")

TOOLKIT_MAP = [
    ("BBBEE Toolkit (RCOGP)_Template_v.1.4.xlsx",        "rcogp", "generic"),
    ("BBBEE Toolkit (RCOGP QSE)_Template_v.1.1.xlsx",    "rcogp", "qse"),
    ("BBBEE Toolkit (ICT Generic)_Template_v.1.4.xlsx",  "ict",   "generic"),
    ("BBBEE Toolkit (ICT QSE)_Template_v.1.1.xlsx",      "ict",   "qse"),
    ("BBBEE Toolkit (Agri Generic)_Master_v.1.0.1.xlsx", "agri",  "generic"),
    ("BBBEE Toolkit (FSC) Template v1.0.xlsx",            "fsc",   "generic"),
]

# Rendering limits per sheet (keeps files sane-sized)
MAX_RENDER_ROWS = 120
MAX_RENDER_COLS = 40

# Pixel sizing
CHAR_PX       = 8      # pixels per Excel character-width unit
PT_PX         = 1.4   # pixels per Excel point (row height)
DEFAULT_COL_W = 8.43  # Excel default column width in chars
DEFAULT_ROW_H = 15.0  # Excel default row height in points
HEADER_H      = 24    # px — sheet name banner at top of each image
PADDING       = 4     # px — inner cell text padding

# Fonts — try a few common Windows/system locations
_FONT_PATHS = [
    "C:/Windows/Fonts/calibri.ttf",
    "C:/Windows/Fonts/arial.ttf",
    "C:/Windows/Fonts/verdana.ttf",
]
_FONT_CACHE: dict = {}

def get_font(size: int = 9):
    if size in _FONT_CACHE:
        return _FONT_CACHE[size]
    for fp in _FONT_PATHS:
        if Path(fp).exists():
            try:
                f = ImageFont.truetype(fp, size)
                _FONT_CACHE[size] = f
                return f
            except Exception:
                pass
    f = ImageFont.load_default()
    _FONT_CACHE[size] = f
    return f

# ---------------------------------------------------------------------------
# Colour helpers
# ---------------------------------------------------------------------------

def _parse_argb(argb: str) -> tuple[int,int,int]:
    """Convert 8-char ARGB hex → (R, G, B)."""
    argb = (argb or "").strip().upper()
    if len(argb) == 6:
        return int(argb[0:2],16), int(argb[2:4],16), int(argb[4:6],16)
    if len(argb) == 8:
        return int(argb[2:4],16), int(argb[4:6],16), int(argb[6:8],16)
    return (255, 255, 255)

# Theme colours (Office default theme — approximate)
_THEME_COLORS = [
    (255,255,255), (0,0,0),   (238,236,225), (31,73,125),
    (79,129,189),  (192,80,77),(155,187,89),  (128,100,162),
    (75,172,198),  (247,150,70),
]

def _theme_rgb(theme_idx: int, tint: float = 0.0) -> tuple[int,int,int]:
    base = _THEME_COLORS[theme_idx % len(_THEME_COLORS)]
    if tint == 0.0:
        return base
    r, g, b = base
    if tint > 0:
        r = int(r + (255-r)*tint)
        g = int(g + (255-g)*tint)
        b = int(b + (255-b)*tint)
    else:
        r = int(r*(1+tint))
        g = int(g*(1+tint))
        b = int(b*(1+tint))
    return (max(0,min(255,r)), max(0,min(255,g)), max(0,min(255,b)))

def cell_bg_color(cell) -> tuple[int,int,int]:
    """Return (R,G,B) background colour for a cell, defaulting to white."""
    try:
        fill = cell.fill
        if fill is None or fill.patternType in (None, "none"):
            return (255, 255, 255)
        # PatternFill: fgColor is the fill colour for "solid"
        color = fill.fgColor
        if color is None:
            return (255, 255, 255)
        ctype = color.type
        if ctype == "rgb":
            argb = color.rgb
            if argb and argb not in ("00000000", "FFFFFFFF", "00FFFFFF"):
                rgb = _parse_argb(argb)
                if rgb != (0,0,0) or fill.patternType == "solid":
                    return rgb
        elif ctype == "theme":
            idx  = getattr(color, "theme", 0) or 0
            tint = getattr(color, "tint", 0.0) or 0.0
            return _theme_rgb(idx, tint)
    except Exception:
        pass
    return (255, 255, 255)

def cell_font_color(cell) -> tuple[int,int,int]:
    try:
        fc = cell.font.color if cell.font else None
        if fc is None:
            return (0,0,0)
        if fc.type == "rgb":
            argb = fc.rgb
            rgb = _parse_argb(argb)
            return rgb
        if fc.type == "theme":
            return _theme_rgb(getattr(fc,"theme",0) or 0, getattr(fc,"tint",0.0) or 0.0)
    except Exception:
        pass
    return (0,0,0)

def is_bold(cell) -> bool:
    try:
        return bool(cell.font and cell.font.bold)
    except Exception:
        return False

# ---------------------------------------------------------------------------
# Sheet dimension helpers
# ---------------------------------------------------------------------------

def sheet_used_area(ws) -> tuple[int,int]:
    """Return (last_row, last_col) of the actually-used data region (1-based)."""
    max_row = ws.max_row or 1
    max_col = ws.max_column or 1
    return (
        min(max_row, MAX_RENDER_ROWS),
        min(max_col, MAX_RENDER_COLS),
    )

def col_px(ws, col_idx: int) -> int:
    """Pixel width for column col_idx (1-based)."""
    letter = get_column_letter(col_idx)
    cd = ws.column_dimensions.get(letter)
    width = (cd.width if cd and cd.width else DEFAULT_COL_W) or DEFAULT_COL_W
    return max(4, int(width * CHAR_PX))

def row_px(ws, row_idx: int) -> int:
    """Pixel height for row row_idx (1-based)."""
    rd = ws.row_dimensions.get(row_idx)
    height = (rd.height if rd and rd.height else DEFAULT_ROW_H) or DEFAULT_ROW_H
    return max(4, int(height * PT_PX))

# ---------------------------------------------------------------------------
# Merged-cell lookup
# ---------------------------------------------------------------------------

def build_merge_map(ws) -> dict:
    """
    Returns a dict: (row,col) → (top_row, left_col, span_rows, span_cols)
    for every cell that is part of a merged range (the top-left cell stores
    the span; all others store None meaning "skip draw").
    """
    merge_info: dict = {}
    for merged_range in ws.merged_cells.ranges:
        min_r = merged_range.min_row
        min_c = merged_range.min_col
        max_r = merged_range.max_row
        max_c = merged_range.max_col
        span_r = max_r - min_r + 1
        span_c = max_c - min_c + 1
        # top-left: store span
        merge_info[(min_r, min_c)] = (min_r, min_c, span_r, span_c)
        # all others in range: mark as covered
        for r in range(min_r, max_r + 1):
            for c in range(min_c, max_c + 1):
                if (r, c) != (min_r, min_c):
                    merge_info[(r, c)] = None
    return merge_info

# ---------------------------------------------------------------------------
# Sheet → PNG
# ---------------------------------------------------------------------------

def render_sheet(ws, sheet_name: str, out_path: Path) -> dict:
    """
    Render worksheet ws to a PNG at out_path.
    Returns a metadata dict for the index.
    """
    last_row, last_col = sheet_used_area(ws)

    # Pre-compute cumulative pixel offsets
    col_widths = [col_px(ws, c) for c in range(1, last_col + 1)]
    row_heights = [row_px(ws, r) for r in range(1, last_row + 1)]

    total_w = sum(col_widths) + 1
    total_h = sum(row_heights) + HEADER_H + 1

    img = Image.new("RGB", (total_w, total_h), (240, 240, 240))
    draw = ImageDraw.Draw(img)

    # Header banner
    draw.rectangle([0, 0, total_w, HEADER_H], fill=(31, 73, 125))
    font_hdr = get_font(11)
    draw.text((6, 4), sheet_name, fill=(255,255,255), font=font_hdr)

    # Build merge map
    try:
        merge_map = build_merge_map(ws)
    except Exception:
        merge_map = {}

    font_sm = get_font(8)
    font_bd = get_font(9)

    # Accumulate stats
    data_cells   = 0
    color_cells  = 0
    has_scoring  = False
    has_input    = False
    has_summary  = False

    # Draw cells
    y = HEADER_H
    for r in range(1, last_row + 1):
        x = 0
        rh = row_heights[r - 1]
        for c in range(1, last_col + 1):
            cw = col_widths[c - 1]
            key = (r, c)
            minfo = merge_map.get(key, "NOT_IN_MAP")

            if minfo is None:
                # Covered by a merge — skip drawing
                x += cw
                continue

            # Determine draw rect (expand for merged cells)
            if isinstance(minfo, tuple):
                _, _, span_r, span_c = minfo
                draw_w = sum(col_widths[c-1 : c-1+span_c]) + span_c - 1
                draw_h = sum(row_heights[r-1 : r-1+span_r]) + span_r - 1
            else:
                draw_w = cw
                draw_h = rh

            try:
                cell = ws.cell(row=r, column=c)
            except Exception:
                x += cw
                continue

            bg = cell_bg_color(cell)
            fg = cell_font_color(cell)
            bold = is_bold(cell)
            val  = cell.value

            # Draw cell background
            draw.rectangle([x, y, x + draw_w, y + draw_h], fill=bg)

            # Thin border
            border_color = (180, 180, 180)
            draw.rectangle([x, y, x + draw_w, y + draw_h],
                           outline=border_color)

            # Cell text
            if val is not None:
                data_cells += 1
                text = str(val)
                # Truncate long text
                max_chars = max(4, draw_w // 5)
                if len(text) > max_chars:
                    text = text[:max_chars - 1] + "…"
                font = font_bd if bold else font_sm
                # Choose contrasting text colour
                lum = 0.299*bg[0] + 0.587*bg[1] + 0.114*bg[2]
                txt_color = (0,0,0) if lum > 128 else (255,255,255)
                draw.text((x + PADDING, y + PADDING), text,
                          fill=txt_color, font=font)

                # Keyword detection
                tl = text.lower()
                if any(k in tl for k in ("score","points","achieved","target","weighting")):
                    has_scoring = True
                if any(k in tl for k in ("input","enter","insert","type")):
                    has_input = True
                if any(k in tl for k in ("total","summary","result","grand")):
                    has_summary = True

            if bg != (255,255,255):
                color_cells += 1

            x += cw
        y += rh

    out_path.parent.mkdir(parents=True, exist_ok=True)
    img.save(str(out_path), "PNG", optimize=True)

    # Observations
    observations = []
    if has_scoring:
        observations.append("scoring table")
    if has_input:
        observations.append("input form")
    if has_summary:
        observations.append("summary/totals")
    if color_cells > data_cells // 4:
        observations.append("colour-coded layout")
    if not observations:
        observations.append("data worksheet")

    return {
        "sheet":        sheet_name,
        "rows":         last_row,
        "cols":         last_col,
        "data_cells":   data_cells,
        "color_cells":  color_cells,
        "observations": observations,
        "png":          str(out_path).replace("\\","/"),
        "img_w":        total_w,
        "img_h":        total_h,
    }

# ---------------------------------------------------------------------------
# Main pipeline
# ---------------------------------------------------------------------------

def sanitize_name(name: str) -> str:
    """Make a sheet name safe for filenames."""
    name = re.sub(r'[\\/:*?"<>|]', "_", name)
    name = re.sub(r'\s+', "_", name.strip())
    return name

def keyword_observations(sheet_name: str) -> list[str]:
    """Extra observations derived from the sheet name alone."""
    nl = sheet_name.lower()
    obs = []
    if "scorecard" in nl:    obs.append("scorecard")
    if "data" in nl:         obs.append("data entry")
    if "toolkit" in nl:      obs.append("toolkit")
    if "calcs" in nl or "calculations" in nl: obs.append("calculations")
    if "empower" in nl:      obs.append("empower slides")
    if "validation" in nl:   obs.append("validation lists")
    if "demographics" in nl: obs.append("employee demographics")
    if "financials" in nl:   obs.append("financial inputs")
    if "report" in nl:       obs.append("report data")
    return obs

def process_workbook(xlsx_path: Path, sector: str, size: str,
                     index_entries: list) -> int:
    rel = xlsx_path.name
    print(f"\n{'='*60}")
    print(f"  {rel}  ->  {sector}/{size}")
    print(f"{'='*60}")

    out_base = OUT_DIR / sector / size
    out_base.mkdir(parents=True, exist_ok=True)

    t0 = time.time()
    print(f"  Loading workbook…", end=" ", flush=True)
    try:
        wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    except Exception as exc:
        print(f"FAILED: {exc}")
        return 0
    print(f"done ({time.time()-t0:.1f}s)  [{len(wb.sheetnames)} sheets]")

    count = 0
    for idx, sheet_name in enumerate(wb.sheetnames, start=1):
        prefix = f"{idx:02d}"
        safe   = sanitize_name(sheet_name)
        fname  = f"{prefix}_{safe}.png"
        out_p  = out_base / fname

        print(f"  [{idx:2d}/{len(wb.sheetnames)}] {sheet_name:<40}", end=" ", flush=True)
        ts = time.time()

        try:
            ws   = wb[sheet_name]
            meta = render_sheet(ws, sheet_name, out_p)
            elapsed = time.time() - ts

            extra_obs = keyword_observations(sheet_name)
            all_obs   = list(dict.fromkeys(meta["observations"] + extra_obs))

            entry = {
                "sector":       sector,
                "size":         size,
                "sheet_name":   sheet_name,
                "filename":     fname,
                "rel_path":     f"screenshots/{sector}/{size}/{fname}",
                "rows":         meta["rows"],
                "cols":         meta["cols"],
                "data_cells":   meta["data_cells"],
                "observations": all_obs,
                "img_w":        meta["img_w"],
                "img_h":        meta["img_h"],
            }
            index_entries.append(entry)
            print(f"OK ({meta['img_w']}x{meta['img_h']}px, {elapsed:.1f}s)")
            count += 1
        except Exception as exc:
            print(f"ERROR: {exc}")

    wb.close()
    print(f"\n  -> {count} screenshots written to {out_base}")
    return count

# ---------------------------------------------------------------------------
# Index generation
# ---------------------------------------------------------------------------

def write_index(index_entries: list) -> None:
    out_path = OUT_DIR / "index.md"

    lines = [
        "# BBBEE Toolkit Screenshots — Index",
        "",
        f"Generated automatically by `scripts/screenshot-toolkits.py`.",
        f"Total exports: **{len(index_entries)}**",
        "",
        "---",
        "",
    ]

    # Group by sector → size
    from collections import defaultdict
    grouped: dict = defaultdict(lambda: defaultdict(list))
    for e in index_entries:
        grouped[e["sector"]][e["size"]].append(e)

    for sector in sorted(grouped):
        lines.append(f"## {sector.upper()}")
        lines.append("")
        for size in sorted(grouped[sector]):
            lines.append(f"### {sector.upper()} / {size}")
            lines.append("")
            lines.append("| # | Sheet | Rows | Cols | Data cells | Observations | Screenshot |")
            lines.append("|---|-------|------|------|------------|--------------|------------|")
            for i, e in enumerate(grouped[sector][size], start=1):
                obs_str  = ", ".join(e["observations"])
                img_link = f"![{e['sheet_name']}]({e['rel_path']})"
                lines.append(
                    f"| {i} | {e['sheet_name']} | {e['rows']} | {e['cols']} "
                    f"| {e['data_cells']} | {obs_str} | {img_link} |"
                )
            lines.append("")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text("\n".join(lines), encoding="utf-8")
    print(f"\nIndex written → {out_path}")

# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main():
    print("BBBEE Toolkit Screenshot Pipeline")
    print(f"Output dir: {OUT_DIR.resolve()}")
    print()

    index_entries: list = []
    total_screenshots = 0

    for filename, sector, size in TOOLKIT_MAP:
        xlsx_path = TOOLKITS_DIR / filename
        if not xlsx_path.exists():
            print(f"  SKIP (not found): {xlsx_path}")
            continue
        n = process_workbook(xlsx_path, sector, size, index_entries)
        total_screenshots += n

    print(f"\n{'='*60}")
    print(f"  DONE — {total_screenshots} screenshots across "
          f"{len(index_entries)} sheet exports")
    print(f"{'='*60}\n")

    write_index(index_entries)

if __name__ == "__main__":
    main()
