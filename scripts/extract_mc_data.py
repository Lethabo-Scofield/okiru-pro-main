"""
Extract MC employee data and procurement data from Lake Trading Excel.
"""
import warnings
warnings.filterwarnings('ignore')
import openpyxl

EXCEL_PATH = r'C:\Users\Administrator\Downloads\BBBEE Toolkit-20260318T172641Z-1-001\BBBEE Toolkit\BBBEE Toolkits\1. RCOGP (Generic)\Lake Trading  Toolkit (RCOGP).xlsx'

print("Loading Excel...")
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True, read_only=True)

print("Sheets:", [s for s in wb.sheetnames if 'MC' in s or 'Procurement' in s or 'ESD' in s])

def get_sheet(name, max_rows=80, max_cols=15):
    if name not in wb.sheetnames:
        print(f"Sheet '{name}' not found")
        return []
    ws = wb[name]
    rows = []
    for row in ws.iter_rows(min_row=1, max_row=max_rows, max_col=max_cols, values_only=True):
        if any(cell is not None for cell in row):
            rows.append(list(row))
    return rows

print("\n=== MC DATA ===")
for row in get_sheet('MC Data', 80, 15):
    print(row)

print("\n=== MC SCORECARD (full) ===")
for row in get_sheet('MC Scorecard', 80, 12):
    print(row)

print("\n=== PROCUREMENT DATA ===")
for row in get_sheet('Procurement Data', 80, 15):
    print(row)

print("\n=== ESD DATA ===")
for row in get_sheet('ESD Data', 40, 15):
    print(row)

print("\n=== SED DATA ===")
for row in get_sheet('SED Data', 30, 15):
    print(row)

wb.close()
print("\nDone!")
