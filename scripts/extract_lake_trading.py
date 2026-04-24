"""
Extract key values from Lake Trading Toolkit Excel for test fixture creation.
Run with: py scripts/extract_lake_trading.py
"""
import warnings
warnings.filterwarnings('ignore')

import openpyxl
import sys

EXCEL_PATH = r'C:\Users\Administrator\Downloads\BBBEE Toolkit-20260318T172641Z-1-001\BBBEE Toolkit\BBBEE Toolkits\1. RCOGP (Generic)\Lake Trading  Toolkit (RCOGP).xlsx'

print("Loading Excel file (this may take 30-60 seconds for a 24MB file)...")
print("File:", EXCEL_PATH)

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True, read_only=True)

print("\nSheets found:", len(wb.sheetnames))
print("First 15:", wb.sheetnames[:15])
print()

# Find Summary Scorecard sheet
summary_sheet = None
for name in wb.sheetnames:
    if 'summary' in name.lower() or ('scorecard' in name.lower() and 'calcs' not in name.lower() and 'data' not in name.lower()):
        print(f"Potential summary sheet: {name}")

# Print first few rows of sheets that look like summaries
print("\n--- Looking for scorecard results ---")
for sheet_name in wb.sheetnames[:5]:
    ws = wb[sheet_name]
    print(f"\nSheet: {sheet_name}")
    for row in ws.iter_rows(min_row=1, max_row=20, max_col=10, values_only=True):
        if any(cell is not None for cell in row):
            print("  ", row)

wb.close()
print("\nDone!")
