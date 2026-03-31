"""
Full extraction of Lake Trading Toolkit for test fixture creation.
Run with: py scripts/extract_lake_trading_full.py
"""
import warnings
warnings.filterwarnings('ignore')
import openpyxl
import json

EXCEL_PATH = r'C:\Users\Administrator\Downloads\BBBEE Toolkit-20260318T172641Z-1-001\BBBEE Toolkit\BBBEE Toolkits\1. RCOGP (Generic)\Lake Trading  Toolkit (RCOGP).xlsx'

print("Loading Excel (read_only, data_only)...")
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True, read_only=True)

def get_sheet_data(sheet_name, max_rows=100, max_cols=15):
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(min_row=1, max_row=max_rows, max_col=max_cols, values_only=True):
        if any(cell is not None for cell in row):
            rows.append(list(row))
    return rows

# ============================================================
# Summary Scorecard - get all pillar scores
# ============================================================
print("\n" + "="*60)
print("SUMMARY SCORECARD")
print("="*60)
summary = get_sheet_data('Summary Scorecard', 60, 12)
for row in summary:
    print(row)

# ============================================================
# Financials
# ============================================================
print("\n" + "="*60)
print("FINANCIALS (key rows)")
print("="*60)
financials = get_sheet_data('Financials', 20, 10)
for row in financials[:15]:
    print(row)

# ============================================================
# Ownership Scorecard
# ============================================================
print("\n" + "="*60)
print("OWNERSHIP SCORECARD")
print("="*60)
own = get_sheet_data('Ownership Scorecard', 30, 10)
for row in own:
    print(row)

# ============================================================
# MC Scorecard
# ============================================================
print("\n" + "="*60)
print("MC SCORECARD")
print("="*60)
mc = get_sheet_data('MC Scorecard', 40, 10)
for row in mc:
    print(row)

# ============================================================
# Skills Scorecard
# ============================================================
print("\n" + "="*60)
print("SKILLS SCORECARD")
print("="*60)
skills_score = get_sheet_data('Skills Scorecard', 30, 10)
for row in skills_score:
    print(row)

# ============================================================
# Procurement Scorecard
# ============================================================
print("\n" + "="*60)
print("PROCUREMENT SCORECARD")
print("="*60)
proc_score = get_sheet_data('Procurement Scorecard', 30, 10)
for row in proc_score:
    print(row)

# ============================================================
# ESD Scorecard
# ============================================================
print("\n" + "="*60)
print("ESD SCORECARD")
print("="*60)
esd_score = get_sheet_data('ESD Scorecard', 30, 10)
for row in esd_score:
    print(row)

# ============================================================
# SED Scorecard
# ============================================================
print("\n" + "="*60)
print("SED SCORECARD")
print("="*60)
sed_score = get_sheet_data('SED Scorecard', 20, 10)
for row in sed_score:
    print(row)

wb.close()
print("\n" + "="*60)
print("EXTRACTION COMPLETE")
print("="*60)
