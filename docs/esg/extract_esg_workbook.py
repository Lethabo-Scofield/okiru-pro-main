"""
Extract ESG workbook: sheet inventory, cell values, formulas, validation, named ranges.
Optional screenshots via Excel COM (Windows) or matplotlib fallback.
"""
from __future__ import annotations

import json
import re
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
XLSX = ROOT / "Okiru_ESG_Toolkit_v1_7_SG_Consumer_LiveData.xlsx"
OUT_EXTRACT = ROOT / "extracted"
OUT_SCREEN = ROOT / "screenshots"
MAX_ROWS = 200
MAX_COLS = 40


def sanitize(name: str) -> str:
    return re.sub(r'[\\/:*?"<>|]', "_", name.strip()).replace(" ", "_")


def cell_to_json(v):
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    if isinstance(v, float):
        if v == int(v) and abs(v) < 1e15:
            return int(v) if abs(v) < 1e12 else v
        return round(v, 10) if abs(v) < 1e10 else v
    return str(v) if not isinstance(v, (int, float, bool)) else v


def extract_sheet_values(ws, data_only: bool) -> dict:
    rows = []
    formulas = []
    for r in range(1, min(ws.max_row + 1, MAX_ROWS + 1)):
        row_cells = {}
        for c in range(1, min(ws.max_column + 1, MAX_COLS + 1)):
            cell = ws.cell(row=r, column=c)
            v = cell.value
            if v is not None:
                coord = f"{get_column_letter(c)}{r}"
                entry = {"coord": coord, "col": c, "value": cell_to_json(v)}
                if not data_only and isinstance(v, str) and v.startswith("="):
                    entry["formula"] = v
                    formulas.append({"coord": coord, "formula": v})
                row_cells[str(c)] = entry
        if row_cells:
            rows.append({"row": r, "cells": row_cells})
    return {"rows": rows, "formulas": formulas, "max_row": ws.max_row, "max_column": ws.max_column}


def extract_validations(ws) -> list:
    out = []
    dv = getattr(ws, "data_validations", None)
    if not dv:
        return out
    for d in dv.dataValidation:
        out.append({
            "type": d.type,
            "operator": d.operator,
            "formula1": d.formula1,
            "formula2": d.formula2,
            "allow_blank": d.allow_blank,
            "showInputMessage": d.showInputMessage,
            "promptTitle": d.promptTitle,
            "prompt": d.prompt,
            "showErrorMessage": d.showErrorMessage,
            "errorTitle": d.errorTitle,
            "error": d.error,
            "sqref": str(d.sqref) if d.sqref else None,
        })
    return out


def extract_named_ranges(wb, sheet_name: str) -> list:
    out = []
    for name, defn in wb.defined_names.items():
        if defn is None:
            continue
        dests = list(defn.destinations)
        for title, ref in dests:
            if title == sheet_name:
                out.append({"name": name, "ref": ref, "comment": defn.comment})
    return out


def try_excel_com_screenshots(sheet_names: list[str]) -> dict[str, str]:
    results = {}
    try:
        import win32com.client  # type: ignore
    except ImportError:
        return {"_error": "win32com not installed"}

    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        wb = excel.Workbooks.Open(str(XLSX.resolve()), ReadOnly=True)
    except Exception as e:
        excel.Quit()
        return {"_error": f"Open failed: {e}"}

    if wb is None:
        excel.Quit()
        return {"_error": "Workbooks.Open returned None"}
    OUT_SCREEN.mkdir(parents=True, exist_ok=True)

    for sn in sheet_names:
        safe = sanitize(sn)
        out_path = OUT_SCREEN / f"{safe}.png"
        try:
            ws = wb.Worksheets(sn)
            ws.Activate()
            # Copy used range as picture
            used = ws.UsedRange
            used.CopyPicture(Appearance=1, Format=2)  # xlScreen, xlBitmap
            # Fallback: export as PDF then skip - CopyPicture needs clipboard
            # Use ExportAsFixedFormat to PDF, then note limitation
            pdf_path = OUT_SCREEN / f"{safe}.pdf"
            ws.ExportAsFixedFormat(0, str(pdf_path.resolve()))
            results[sn] = f"pdf:{pdf_path.name}"
        except Exception as e:
            results[sn] = f"error:{e}"

    wb.Close(False)
    excel.Quit()
    return results


def try_matplotlib_screenshot(ws, out_path: Path) -> bool:
    try:
        from PIL import Image, ImageDraw, ImageFont
    except ImportError:
        return False

    font = None
    try:
        font = ImageFont.truetype("consola.ttf", 11)
    except Exception:
        try:
            font = ImageFont.truetype("arial.ttf", 11)
        except Exception:
            font = ImageFont.load_default()

    rows = []
    for r in range(1, min(ws.max_row + 1, 80)):
        cells = []
        for c in range(1, min(ws.max_column + 1, 16)):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                s = str(v)
                if len(s) > 28:
                    s = s[:25] + "..."
                cells.append(s)
            else:
                cells.append("")
        if any(cells):
            rows.append(cells)

    if not rows:
        return False

    col_count = max(len(r) for r in rows)
    col_w = 140
    row_h = 18
    pad = 8
    w = col_count * col_w + pad * 2
    h = len(rows) * row_h + pad * 2 + 24
    img = Image.new("RGB", (w, h), (8, 14, 20))
    draw = ImageDraw.Draw(img)
    draw.text((pad, 4), ws.title[:60], fill=(29, 233, 160), font=font)
    y = pad + 20
    for row in rows:
        x = pad
        for cell in row:
            draw.rectangle([x, y, x + col_w - 2, y + row_h - 2], outline=(40, 50, 60))
            draw.text((x + 4, y + 2), cell, fill=(240, 237, 232), font=font)
            x += col_w
        y += row_h
    img.save(out_path)
    return True


def main():
    if not XLSX.exists():
        print(f"Missing workbook: {XLSX}", file=sys.stderr)
        sys.exit(1)

    OUT_EXTRACT.mkdir(parents=True, exist_ok=True)
    OUT_SCREEN.mkdir(parents=True, exist_ok=True)

    print("Loading workbook (values)...", file=sys.stderr)
    wb_data = openpyxl.load_workbook(str(XLSX), data_only=True)
    print("Loading workbook (formulas)...", file=sys.stderr)
    wb_formula = openpyxl.load_workbook(str(XLSX), data_only=False)

    inventory = {
        "file": XLSX.name,
        "sheet_count": len(wb_data.sheetnames),
        "sheet_names": wb_data.sheetnames,
        "defined_names": [],
    }

    for name, defn in wb_formula.defined_names.items():
        if defn is None:
            continue
        dests = [{"sheet": t, "ref": r} for t, r in defn.destinations]
        inventory["defined_names"].append({
            "name": name,
            "destinations": dests,
            "comment": defn.comment,
        })

    screenshot_status = {}
    summary_path = ROOT / "workbook_inventory.json"
    per_sheet_meta = []

    for sn in wb_data.sheetnames:
        safe = sanitize(sn)
        print(f"  Processing: {sn}", file=sys.stderr)
        ws_d = wb_data[sn]
        ws_f = wb_formula[sn]

        val_block = extract_sheet_values(ws_d, data_only=True)
        form_block = extract_sheet_values(ws_f, data_only=False)
        # merge formulas into value block
        formula_map = {f["coord"]: f["formula"] for f in form_block["formulas"]}
        for row in val_block["rows"]:
            for ckey, cell in row["cells"].items():
                coord = cell["coord"]
                if coord in formula_map:
                    cell["formula"] = formula_map[coord]

        validations = extract_validations(ws_f)
        named = extract_named_ranges(wb_formula, sn)

        sheet_doc = {
            "sheet_name": sn,
            "sanitized_name": safe,
            "dimensions": {"max_row": ws_d.max_row, "max_column": ws_d.max_column},
            "named_ranges": named,
            "data_validations": validations,
            "formula_count": len(form_block["formulas"]),
            "rows": val_block["rows"],
        }

        out_json = OUT_EXTRACT / f"{safe}.json"
        with open(out_json, "w", encoding="utf-8") as f:
            json.dump(sheet_doc, f, indent=2, ensure_ascii=False, default=str)

        # compact markdown summary
        md_lines = [f"# {sn}", "", f"- Rows scanned: up to {MAX_ROWS}", f"- Formulas: {len(form_block['formulas'])}", ""]
        if validations:
            md_lines.append("## Data validations")
            for v in validations[:20]:
                md_lines.append(f"- `{v.get('sqref')}` type={v.get('type')} formula1={v.get('formula1')}")
            md_lines.append("")
        if named:
            md_lines.append("## Named ranges")
            for n in named:
                md_lines.append(f"- `{n['name']}` → {n['ref']}")
            md_lines.append("")
        md_lines.append("## Sample rows (first 30)")
        for row in val_block["rows"][:30]:
            parts = []
            for ck in sorted(row["cells"], key=lambda x: int(x)):
                c = row["cells"][ck]
                v = c.get("value")
                if c.get("formula"):
                    parts.append(f"{c['coord']}={v} [{c['formula']}]")
                else:
                    parts.append(f"{c['coord']}={v}")
            if parts:
                md_lines.append(f"R{row['row']}: " + " | ".join(parts[:8]))

        out_md = OUT_EXTRACT / f"{safe}.md"
        out_md.write_text("\n".join(md_lines), encoding="utf-8")

        # matplotlib screenshot
        png_path = OUT_SCREEN / f"{safe}.png"
        ok = try_matplotlib_screenshot(ws_d, png_path)
        screenshot_status[sn] = "png_ok" if ok else "png_failed"

        per_sheet_meta.append({
            "name": sn,
            "safe": safe,
            "rows": ws_d.max_row,
            "cols": ws_d.max_column,
            "formulas": len(form_block["formulas"]),
            "validations": len(validations),
            "screenshot": screenshot_status[sn],
        })

    wb_data.close()
    wb_formula.close()

    # Try Excel COM for better exports (optional) — skip if workbook already closed
    com_results = try_excel_com_screenshots(list(screenshot_status.keys()))
    if com_results and "_error" not in com_results:
        for sn, status in com_results.items():
            if sn in screenshot_status and status.startswith("pdf:"):
                screenshot_status[sn] = status

    inventory["sheets"] = per_sheet_meta
    inventory["screenshots"] = screenshot_status
    with open(summary_path, "w", encoding="utf-8") as f:
        json.dump(inventory, f, indent=2, ensure_ascii=False)

    print(json.dumps(inventory, indent=2))
    print(f"\nWrote inventory to {summary_path}", file=sys.stderr)


if __name__ == "__main__":
    main()
