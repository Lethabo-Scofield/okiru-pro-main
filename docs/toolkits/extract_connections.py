"""
extract_connections.py  —  VERBATIM toolkit connection extractor (v2, memory-safe).

extract_fast.py used data_only=True and captured only cached *values*. This
captures the *links*: formulas, the cross-sheet reference graph, data-validation
dropdowns (the only viable values a user may enter), defined names and merges.

Design notes (learned the hard way):
  * openpyxl read_only=False on these 24-94 MB workbooks hangs / OOMs. We use
    read_only=True for cells (formulas with data_only=False; cached values with
    data_only=True) — fast (~3 s for 52 sheets).
  * read_only mode does NOT expose data validations, and z.read() of a whole
    sheet XML can OOM (one sheet decompresses to GBs of styled blanks). So DVs
    and merges are scanned from the zip with a chunked, bounded regex stream.

Outputs per toolkit:
  docs/toolkits/extracted_formulas/<SECTOR>.json   (machine-readable)
  docs/toolkits/extracted_formulas/<SECTOR>.md     (human/agent digest)

Run:  py -3 docs/toolkits/extract_connections.py [SECTOR]
"""
import json
import re
import sys
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

import openpyxl
from openpyxl.utils import range_boundaries

TOOLKIT_DIR = Path(__file__).parent
OUT_DIR = TOOLKIT_DIR / "extracted_formulas"
OUT_DIR.mkdir(exist_ok=True)

TOOLKITS = [
    ("RCOGP_Generic", "BBBEE Toolkit (RCOGP)_Template_v.1.4.xlsx"),
    ("RCOGP_QSE", "BBBEE Toolkit (RCOGP QSE)_Template_v.1.1.xlsx"),
    ("ICT_Generic", "BBBEE Toolkit (ICT Generic)_Template_v.1.4.xlsx"),
    ("ICT_QSE", "BBBEE Toolkit (ICT QSE)_Template_v.1.1.xlsx"),
    ("FSC_Generic", "BBBEE Toolkit (FSC) Template v1.0.xlsx"),
    ("AGRI_Generic", "BBBEE Toolkit (Agri Generic)_Master_v.1.0.1.xlsx"),
]

FORMULA_SHEET_KW = [
    "scorecard", "summary", "eap", "industry norm", "client information",
    "financ", "calc", "yes", "scoring scale", "level", "ef &", "afs",
    "sed & ce", "ee scorecard", "deemed", "norms", "validation",
]
DATA_SHEET_KW = ["data", "register", "fleet", "employees", "training", "report data", "tmps"]

MAX_FORMULA_ROWS = 300
MAX_DATA_ROWS = 14
MAX_COLS = 45
CROSS_SHEET_RE = re.compile(r"'?([A-Za-z0-9 _&()\-\.]+?)'?!\$?[A-Z]{1,3}\$?\d+")


def classify(sn):
    s = sn.lower().strip()
    is_data = any(k in s for k in DATA_SHEET_KW)
    if is_data and not s.endswith("scorecard"):
        return "data"
    if any(k in s for k in FORMULA_SHEET_KW):
        return "formula"
    return "other"


# ---------------------------------------------------------------------------
# Sheet name -> worksheet xml path (parse workbook.xml + rels)
# ---------------------------------------------------------------------------
def sheet_xml_map(zf):
    wb_xml = zf.read("xl/workbook.xml").decode("utf-8", "ignore")
    rels = zf.read("xl/_rels/workbook.xml.rels").decode("utf-8", "ignore")
    rid_to_target = dict(re.findall(r'Id="([^"]+)"[^>]*Target="([^"]+)"', rels))
    name_to_path = {}
    for m in re.finditer(r'<sheet [^>]*name="([^"]+)"[^>]*r:id="([^"]+)"', wb_xml):
        name, rid = m.group(1), m.group(2)
        tgt = rid_to_target.get(rid, "")
        if not tgt:
            continue
        path = tgt if tgt.startswith("xl/") else ("xl/" + tgt.lstrip("/"))
        path = path.replace("/worksheets/../", "/")
        name_to_path[name] = path
    return name_to_path


# ---------------------------------------------------------------------------
# Bounded streaming scan of one worksheet XML for dataValidations + mergeCells.
# ---------------------------------------------------------------------------
DV_RE = re.compile(
    r"<(?:x14:)?dataValidation\b[^>]*?(?:/>|>.*?</(?:x14:)?dataValidation>)", re.S
)
MERGE_RE = re.compile(r'<mergeCell ref="([^"]+)"')


def parse_dv(blob):
    typ = (re.search(r'\btype="([^"]+)"', blob) or [None, None])[1]
    sqref = (re.search(r'\bsqref="([^"]+)"', blob) or [None, None])[1]
    if sqref is None:
        m = re.search(r"<xm:sqref>(.*?)</xm:sqref>", blob, re.S)
        sqref = m.group(1).strip() if m else None
    f1 = re.search(r"<formula1>(.*?)</formula1>", blob, re.S)
    if not f1:
        f1 = re.search(r"<xm:f>(.*?)</xm:f>", blob, re.S)
    formula1 = f1.group(1).strip() if f1 else (
        (re.search(r'\bformula1="([^"]*)"', blob) or [None, None])[1]
    )
    return {"type": typ, "sqref": sqref, "formula1": formula1}


def scan_sheet_xml(zf, path):
    dvs, merges = [], []
    seen_dv, seen_mg = set(), set()
    tail = ""
    try:
        with zf.open(path) as f:
            while True:
                chunk = f.read(1 << 20)
                if not chunk:
                    break
                text = tail + chunk.decode("utf-8", "ignore")
                for m in MERGE_RE.finditer(text):
                    if m.group(1) not in seen_mg:
                        seen_mg.add(m.group(1))
                        merges.append(m.group(1))
                for m in DV_RE.finditer(text):
                    blob = m.group(0)
                    if blob not in seen_dv:
                        seen_dv.add(blob)
                        dvs.append(parse_dv(blob))
                tail = text[-8192:]  # overlap so boundary-split tags still match
    except Exception as e:
        return [], [], str(e)
    return dvs, merges, None


# ---------------------------------------------------------------------------
# Resolve a data-validation list source to concrete options.
# ---------------------------------------------------------------------------
def resolve_list(formula1, wb_v, defined=None, _depth=0):
    defined = defined or {}
    if not formula1 or _depth > 3:
        return None
    s = str(formula1).strip()
    if not s.startswith("="):
        s2 = s.strip('"')
        # A bare token may be a NAMED RANGE (e.g. "Designations", "Races", "Sizes").
        if "!" not in s2 and "$" not in s2 and s2 and "," not in s2 and s2 in defined:
            return resolve_list("=" + str(defined[s2]).lstrip("="), wb_v, defined, _depth + 1)
        if "!" not in s2 and "$" not in s2 and s2:
            return {"options": [p.strip() for p in s2.split(",") if p.strip()]}
    ref = s.lstrip("=")
    # Named range with no sheet qualifier.
    if "!" not in ref and ref in defined:
        return resolve_list("=" + str(defined[ref]).lstrip("="), wb_v, defined, _depth + 1)
    sheet, rng = None, ref
    if "!" in ref:
        sheet, rng = ref.split("!", 1)
        sheet = sheet.strip("'")
    if sheet is None or sheet not in wb_v.sheetnames:
        return {"source": ref}
    try:
        c1, r1, c2, r2 = range_boundaries(rng.replace("$", ""))
    except Exception:
        return {"source": ref}
    r1, r2 = r1 or 1, r2 or r1 or 1
    c1, c2 = c1 or 1, c2 or c1 or 1
    if (r2 - r1 + 1) * (c2 - c1 + 1) > 600:
        return {"source": ref, "note": "range too large to resolve"}
    vals = []
    try:
        ws = wb_v[sheet]
        for row in ws.iter_rows(min_row=r1, max_row=r2, min_col=c1, max_col=c2):
            for cell in row:
                if cell.value not in (None, ""):
                    vals.append(cell.value)
    except Exception:
        return {"source": ref}
    return {"source": ref, "options": vals} if vals else {"source": ref}


# ---------------------------------------------------------------------------
# Cell extraction (formulas + cached values) via two read_only loads.
# ---------------------------------------------------------------------------
def extract_cells(ws_f, ws_v, kind):
    cap = MAX_FORMULA_ROWS if kind in ("formula", "other") else MAX_DATA_ROWS
    rows_f = ws_f.iter_rows(min_row=1, max_row=cap, max_col=MAX_COLS)
    rows_v = ws_v.iter_rows(min_row=1, max_row=cap, max_col=MAX_COLS)
    cells = []
    for rf, rv in zip(rows_f, rows_v):
        for cf, cv in zip(rf, rv):
            f = cf.value
            v = cv.value
            if f is None and v is None:
                continue
            is_formula = isinstance(f, str) and f.startswith("=")
            rec = {"ref": cf.coordinate}
            if is_formula:
                rec["formula"] = f
                rec["value"] = v
                xr = sorted({m.strip() for m in CROSS_SHEET_RE.findall(f)})
                if xr:
                    rec["xrefs"] = xr
            else:
                rec["value"] = f if f is not None else v
            cells.append(rec)
    return cells


def process(name, filename):
    fp = TOOLKIT_DIR / filename
    if not fp.exists():
        return {"error": f"not found: {fp}"}
    print(f"  {name}: loading formulas...", file=sys.stderr)
    wb_f = openpyxl.load_workbook(str(fp), data_only=False, read_only=True)
    print(f"  {name}: loading values...", file=sys.stderr)
    wb_v = openpyxl.load_workbook(str(fp), data_only=True, read_only=True)
    zf = zipfile.ZipFile(str(fp))
    xmlmap = sheet_xml_map(zf)

    defined = {}
    try:
        for nm in wb_f.defined_names.keys():
            if nm.startswith("_xlchart") or nm.startswith("_xlfn"):
                continue
            dn = wb_f.defined_names[nm]
            defined[nm] = getattr(dn, "value", str(dn))
    except Exception as e:
        defined["_error"] = str(e)

    result = {
        "file": filename,
        "sheet_count": len(wb_f.sheetnames),
        "all_sheets": wb_f.sheetnames,
        "defined_names": defined,
        "sheets": {},
    }

    for sn in wb_f.sheetnames:
        kind = classify(sn)
        print(f"    {sn} [{kind}]", file=sys.stderr)
        ws_f = wb_f[sn]
        ws_v = wb_v[sn] if sn in wb_v.sheetnames else ws_f
        cells = extract_cells(ws_f, ws_v, kind)
        dvs, merges, dverr = ([], [], None)
        path = xmlmap.get(sn)
        if path:
            dvs, merges, dverr = scan_sheet_xml(zf, path)
        for dv in dvs:
            if dv.get("type") == "list":
                dv["resolved"] = resolve_list(dv.get("formula1"), wb_v, defined)
        result["sheets"][sn] = {
            "kind": kind,
            "cells": cells,
            "validations": dvs,
            "merged": merges,
            "dv_error": dverr,
        }

    wb_f.close()
    wb_v.close()
    zf.close()
    return result


def write_md(name, data):
    L = [f"# {name} — Toolkit Connection Extract", ""]
    L += [f"**File:** `{data['file']}` · **Sheets:** {data['sheet_count']}", ""]
    L += ["**All sheets:** " + ", ".join(f"`{s}`" for s in data["all_sheets"]), ""]

    dn = {k: v for k, v in data.get("defined_names", {}).items() if not k.startswith("_")}
    if dn:
        L += ["## Defined names (named ranges)", ""]
        for nm, v in sorted(dn.items()):
            L.append(f"- `{nm}` → `{v}`")
        L.append("")

    L += ["## Formula sheets — verbatim formulas", ""]
    for sn, sd in data["sheets"].items():
        if sd["kind"] != "formula":
            continue
        fcells = [c for c in sd["cells"] if "formula" in c]
        labels = [c for c in sd["cells"] if "formula" not in c]
        L += [f"### `{sn}`  ({len(fcells)} formulas, {len(labels)} labels)", ""]
        if fcells:
            L += ["| Cell | Formula | Cached | Cross-refs |",
                  "|------|---------|--------|------------|"]
            for c in fcells:
                fm = str(c["formula"]).replace("|", "\\|")
                vv = str(c.get("value", "")).replace("|", "\\|")[:32]
                xr = ", ".join(c.get("xrefs", []))[:64]
                L.append(f"| {c['ref']} | `{fm[:150]}` | {vv} | {xr} |")
            L.append("")
        if labels:
            L += ["<details><summary>Labels / constants</summary>", ""]
            for c in labels:
                L.append(f"- `{c['ref']}`: {str(c.get('value','')).replace(chr(10),' ')[:140]}")
            L += ["", "</details>", ""]

    L += ["## Input options — data validations / dropdowns", ""]
    any_dv = False
    for sn, sd in data["sheets"].items():
        dvs = [d for d in sd["validations"] if d.get("type") == "list" or d.get("formula1")]
        if not dvs:
            continue
        any_dv = True
        L += [f"### `{sn}`", "", "| Cells (sqref) | Type | Options / source |",
              "|---------------|------|------------------|"]
        for dv in sd["validations"]:
            opts = ""
            res = dv.get("resolved")
            if isinstance(res, dict):
                if res.get("options"):
                    opts = ", ".join(str(o) for o in res["options"])
                else:
                    opts = f"(source: {res.get('source','?')}) {res.get('note','')}"
            elif dv.get("formula1"):
                opts = str(dv["formula1"])
            opts = opts.replace("|", "\\|").replace("\n", " ")[:140]
            L.append(f"| {dv.get('sqref')} | {dv.get('type')} | {opts} |")
        L.append("")
    if not any_dv:
        L += ["_No data validations found._", ""]

    L += ["## Data-entry sheet headers", ""]
    for sn, sd in data["sheets"].items():
        if sd["kind"] != "data":
            continue
        hdr = [c for c in sd["cells"] if "formula" not in c and c.get("value") not in (None, "")]
        if not hdr:
            continue
        L += [f"### `{sn}`", ""]
        for c in hdr[:80]:
            L.append(f"- `{c['ref']}`: {str(c.get('value',''))[:100]}")
        L.append("")

    return "\n".join(L)


def main():
    target = sys.argv[1] if len(sys.argv) > 1 else None
    for name, filename in TOOLKITS:
        if target and target != name:
            continue
        data = process(name, filename)
        (OUT_DIR / f"{name}.json").write_text(
            json.dumps(data, indent=1, default=str, ensure_ascii=False), encoding="utf-8"
        )
        if "error" not in data:
            (OUT_DIR / f"{name}.md").write_text(write_md(name, data), encoding="utf-8")
        print(f"  wrote {name}.json / .md", file=sys.stderr)
    print("done", file=sys.stderr)


if __name__ == "__main__":
    main()
