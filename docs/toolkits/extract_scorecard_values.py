"""
Extract all scorecard configuration values from the 6 B-BBEE Excel toolkit files.
Outputs a structured JSON for each sector with every pillar, criterion, target, max points,
industry norms, category weightings, benefit factors, level thresholds, and EAP targets.
"""
import openpyxl
import json
import os
import sys
import re
from pathlib import Path

TOOLKIT_DIR = Path(__file__).parent

TOOLKITS = {
    "RCOGP_Generic": "BBBEE Toolkit (RCOGP)_Template_v.1.4.xlsx",
    "ICT_Generic": "BBBEE Toolkit (ICT Generic)_Template_v.1.4.xlsx",
    "ICT_QSE": "BBBEE Toolkit (ICT QSE)_Template_v.1.1.xlsx",
    "RCOGP_QSE": "BBBEE Toolkit (RCOGP QSE)_Template_v.1.1.xlsx",
    "FSC_Generic": "BBBEE Toolkit (FSC) Template v1.0.xlsx",
    "AGRI_Generic": "BBBEE Toolkit (Agri Generic)_Master_v.1.0.1.xlsx",
}


def safe_val(ws, row, col):
    """Get cell value safely."""
    try:
        v = ws.cell(row=row, column=col).value
        return v
    except Exception:
        return None


def find_sheet(wb, *candidates):
    """Find sheet by trying multiple name candidates (case-insensitive partial match)."""
    for name in wb.sheetnames:
        lower = name.lower().strip()
        for c in candidates:
            if c.lower() in lower:
                return wb[name]
    return None


def extract_summary_scorecard(wb):
    """Extract pillar max points and scores from Summary Scorecard sheet."""
    ws = find_sheet(wb, "summary scorecard", "summary")
    if not ws:
        return {"error": "Summary Scorecard sheet not found", "available_sheets": wb.sheetnames[:20]}

    pillars = {}
    for row in range(1, ws.max_row + 1):
        for col in range(1, min(ws.max_column + 1, 15)):
            val = safe_val(ws, row, col)
            if isinstance(val, str):
                val_lower = val.strip().lower()
                if any(k in val_lower for k in ["ownership", "management", "skills", "procurement",
                                                  "supplier dev", "enterprise dev", "socio", "sed",
                                                  "yes", "empowerment financing", "access to financial",
                                                  "consumer education", "grand total", "total"]):
                    row_data = {}
                    for c2 in range(col, min(col + 10, ws.max_column + 1)):
                        v2 = safe_val(ws, row, c2)
                        if v2 is not None:
                            row_data[f"col_{c2}"] = v2
                    header_vals = []
                    for c2 in range(1, min(ws.max_column + 1, 15)):
                        hv = safe_val(ws, max(1, row - 3), c2)
                        if hv is not None:
                            header_vals.append(str(hv))
                    pillars[val.strip()] = {
                        "row": row,
                        "data": row_data,
                    }
    return pillars


def extract_pillar_scorecard(wb, pillar_name, search_terms):
    """Extract criteria from a pillar scorecard sheet."""
    ws = find_sheet(wb, *search_terms)
    if not ws:
        return {"error": f"{pillar_name} scorecard sheet not found"}

    criteria = []
    header_row = None
    headers = {}

    for row in range(1, min(ws.max_row + 1, 100)):
        for col in range(1, min(ws.max_column + 1, 20)):
            val = safe_val(ws, row, col)
            if isinstance(val, str):
                vl = val.strip().lower()
                if any(k in vl for k in ["indicator", "criteria", "element", "description"]):
                    header_row = row
                    for c2 in range(1, min(ws.max_column + 1, 20)):
                        hv = safe_val(ws, row, c2)
                        if isinstance(hv, str):
                            headers[c2] = hv.strip()
                    break
        if header_row:
            break

    if header_row:
        for row in range(header_row + 1, min(ws.max_row + 1, header_row + 50)):
            row_data = {}
            has_content = False
            for col, header in headers.items():
                val = safe_val(ws, row, col)
                if val is not None:
                    has_content = True
                    row_data[header] = val
            for col in range(1, min(ws.max_column + 1, 20)):
                if col not in headers:
                    val = safe_val(ws, row, col)
                    if val is not None:
                        row_data[f"col_{col}"] = val
            if has_content and row_data:
                criteria.append(row_data)

    return {
        "sheet_name": ws.title,
        "header_row": header_row,
        "headers": headers,
        "criteria": criteria,
    }


def extract_industry_norms(wb):
    """Extract Industry Norms lookup table."""
    ws = find_sheet(wb, "industry norm", "norms")
    if not ws:
        return {"error": "Industry Norms sheet not found"}

    norms = []
    for row in range(1, min(ws.max_row + 1, 200)):
        row_data = {}
        for col in range(1, min(ws.max_column + 1, 15)):
            val = safe_val(ws, row, col)
            if val is not None:
                row_data[f"col_{col}"] = val
        if row_data:
            norms.append({"row": row, "data": row_data})
    return {"sheet_name": ws.title if ws else "N/A", "rows": norms}


def extract_eap(wb):
    """Extract EAP (Economically Active Population) targets."""
    ws = find_sheet(wb, "eap")
    if not ws:
        return {"error": "EAP sheet not found (normal for QSE)"}

    data = []
    for row in range(1, min(ws.max_row + 1, 100)):
        row_data = {}
        for col in range(1, min(ws.max_column + 1, 20)):
            val = safe_val(ws, row, col)
            if val is not None:
                row_data[f"col_{col}"] = val
        if row_data:
            data.append({"row": row, "data": row_data})
    return {"sheet_name": ws.title, "rows": data}


def extract_level_thresholds(wb):
    """Extract level determination / scoring scale."""
    ws = find_sheet(wb, "scoring scale", "level", "scorecard calc")
    if not ws:
        return {"error": "Level thresholds sheet not found"}

    thresholds = []
    for row in range(1, min(ws.max_row + 1, 100)):
        for col in range(1, min(ws.max_column + 1, 15)):
            val = safe_val(ws, row, col)
            if isinstance(val, (int, float)) and 1 <= val <= 8:
                next_vals = []
                for c2 in range(col + 1, min(col + 5, ws.max_column + 1)):
                    nv = safe_val(ws, row, c2)
                    if nv is not None:
                        next_vals.append(nv)
                if next_vals:
                    thresholds.append({"level": val, "adjacent_values": next_vals, "row": row})
    return thresholds


def extract_category_weightings(wb):
    """Extract skills category weightings (A-G)."""
    ws = find_sheet(wb, "skills calc", "skills scorecard")
    if not ws:
        return {"error": "Skills calcs sheet not found"}

    categories = []
    for row in range(1, min(ws.max_row + 1, 200)):
        for col in range(1, min(ws.max_column + 1, 20)):
            val = safe_val(ws, row, col)
            if isinstance(val, str) and val.strip() in ["A", "B", "C", "D", "E", "F", "G"]:
                row_data = {"category": val.strip(), "row": row}
                for c2 in range(1, min(ws.max_column + 1, 20)):
                    v2 = safe_val(ws, row, c2)
                    if v2 is not None:
                        row_data[f"col_{c2}"] = v2
                categories.append(row_data)
    return {"sheet_name": ws.title, "categories": categories}


def extract_all_sheet_names(wb):
    """Get all sheet names for reference."""
    return wb.sheetnames


def scan_for_targets_and_maxpoints(wb):
    """Scan all scorecard sheets for target% and max points columns."""
    results = {}
    scorecard_sheets = [s for s in wb.sheetnames
                        if "scorecard" in s.lower() and "summary" not in s.lower()]

    for sheet_name in scorecard_sheets:
        ws = wb[sheet_name]
        sheet_data = {"rows": []}

        for row in range(1, min(ws.max_row + 1, 80)):
            row_vals = {}
            for col in range(1, min(ws.max_column + 1, 20)):
                val = safe_val(ws, row, col)
                if val is not None:
                    row_vals[f"col_{col}"] = val
            if row_vals:
                sheet_data["rows"].append({"row": row, "data": row_vals})

        results[sheet_name] = sheet_data

    return results


def extract_benefit_factors(wb):
    """Extract ESD benefit factor table."""
    ws = find_sheet(wb, "esd", "benefit", "contribution")
    if not ws:
        return {"error": "Benefit factors sheet not found"}

    factors = []
    for row in range(1, min(ws.max_row + 1, 100)):
        for col in range(1, min(ws.max_column + 1, 15)):
            val = safe_val(ws, row, col)
            if isinstance(val, str) and any(k in val.lower() for k in
                                             ["grant", "loan", "guarantee", "discount",
                                              "direct cost", "overhead", "professional",
                                              "employee time", "shorter", "equity",
                                              "minority", "interest"]):
                row_data = {"type": val.strip(), "row": row}
                for c2 in range(1, min(ws.max_column + 1, 15)):
                    v2 = safe_val(ws, row, c2)
                    if v2 is not None:
                        row_data[f"col_{c2}"] = v2
                factors.append(row_data)
    return {"factors": factors}


def process_toolkit(name, filename):
    """Process a single toolkit file."""
    filepath = TOOLKIT_DIR / filename
    if not filepath.exists():
        return {"error": f"File not found: {filepath}"}

    print(f"\n{'='*60}", file=sys.stderr)
    print(f"Processing: {name} ({filename})", file=sys.stderr)
    print(f"{'='*60}", file=sys.stderr)

    try:
        wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)
    except Exception as e:
        return {"error": f"Failed to open: {e}"}

    print(f"  Sheets: {len(wb.sheetnames)}", file=sys.stderr)

    result = {
        "file": filename,
        "sheet_count": len(wb.sheetnames),
        "all_sheets": extract_all_sheet_names(wb),
        "summary_scorecard": extract_summary_scorecard(wb),
        "scorecard_sheets": scan_for_targets_and_maxpoints(wb),
        "industry_norms": extract_industry_norms(wb),
        "eap": extract_eap(wb),
        "category_weightings": extract_category_weightings(wb),
        "benefit_factors": extract_benefit_factors(wb),
    }

    wb.close()
    print(f"  Done.", file=sys.stderr)
    return result


def main():
    all_results = {}
    for name, filename in TOOLKITS.items():
        all_results[name] = process_toolkit(name, filename)

    output_path = TOOLKIT_DIR / "extracted_values.json"
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(all_results, f, indent=2, default=str, ensure_ascii=False)

    print(f"\nResults written to: {output_path}", file=sys.stderr)
    print(json.dumps({"status": "complete", "file": str(output_path)}, indent=2))


if __name__ == "__main__":
    main()
