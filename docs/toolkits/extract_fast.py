"""
Fast targeted extraction of scorecard values from B-BBEE toolkit Excel files.
Only reads scorecard sheets (small sheets) and industry norms.
Avoids scanning large data sheets.
"""
import openpyxl
import json
import sys
from pathlib import Path

TOOLKIT_DIR = Path(__file__).parent

TOOLKITS = [
    ("RCOGP_Generic", "BBBEE Toolkit (RCOGP)_Template_v.1.4.xlsx"),
    ("ICT_Generic", "BBBEE Toolkit (ICT Generic)_Template_v.1.4.xlsx"),
    ("ICT_QSE", "BBBEE Toolkit (ICT QSE)_Template_v.1.1.xlsx"),
    ("RCOGP_QSE", "BBBEE Toolkit (RCOGP QSE)_Template_v.1.1.xlsx"),
    ("FSC_Generic", "BBBEE Toolkit (FSC) Template v1.0.xlsx"),
    ("AGRI_Generic", "BBBEE Toolkit (Agri Generic)_Master_v.1.0.1.xlsx"),
]

TARGET_SHEET_KEYWORDS = [
    "summary scorecard", "ownership scorecard", "mc scorecard",
    "skills scorecard", "procurement scorecard", "esd scorecard",
    "sed scorecard", "yes", "industry norm", "eap",
    "scoring scale", "scorecard calc",
    "ef &", "afs scorecard", "sed & ce",
    "ee scorecard",
]


def val(ws, r, c):
    try:
        return ws.cell(row=r, column=c).value
    except:
        return None


def should_read(sheet_name):
    sn = sheet_name.lower().strip()
    return any(kw in sn for kw in TARGET_SHEET_KEYWORDS)


def read_sheet_block(ws, max_rows=80, max_cols=20):
    rows = []
    for r in range(1, min(ws.max_row + 1, max_rows)):
        row_data = {}
        for c in range(1, min(ws.max_column + 1, max_cols)):
            v = val(ws, r, c)
            if v is not None:
                row_data[c] = v
        if row_data:
            rows.append({"row": r, "cells": row_data})
    return rows


def process_file(name, filename):
    fp = TOOLKIT_DIR / filename
    if not fp.exists():
        return {"error": f"Not found: {fp}"}

    print(f"  Opening {name}...", file=sys.stderr)
    wb = openpyxl.load_workbook(str(fp), read_only=True, data_only=True)
    print(f"  Sheets: {len(wb.sheetnames)}", file=sys.stderr)

    result = {
        "file": filename,
        "sheet_count": len(wb.sheetnames),
        "all_sheets": wb.sheetnames,
        "extracted_sheets": {},
    }

    for sn in wb.sheetnames:
        if should_read(sn):
            print(f"    Reading: {sn}", file=sys.stderr)
            ws = wb[sn]
            result["extracted_sheets"][sn] = read_sheet_block(ws)

    wb.close()
    print(f"  Done: {name}", file=sys.stderr)
    return result


def main():
    target = sys.argv[1] if len(sys.argv) > 1 else None
    results = {}

    for name, filename in TOOLKITS:
        if target and target != name:
            continue
        results[name] = process_file(name, filename)

    out = TOOLKIT_DIR / ("extracted_values.json" if not target else f"extracted_{target}.json")
    with open(out, "w", encoding="utf-8") as f:
        json.dump(results, f, indent=2, default=str, ensure_ascii=False)
    print(f"\nWrote: {out}", file=sys.stderr)


if __name__ == "__main__":
    main()
