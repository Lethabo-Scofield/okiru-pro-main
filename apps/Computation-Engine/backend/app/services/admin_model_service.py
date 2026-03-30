"""Admin model management service.

This module provides the upload/compile/evaluate lifecycle for scorecard models.

The workflow is:
  1. Upload an Excel file (upload endpoint returns immediately).
  2. Background compilation parses + builds a dependency graph, then persists an
     immutable compiled model artifact to ArangoDB.
  3. Evaluation uses the stored artifact and computes values via graph traversal.

The goal is to _compile once_ and then evaluate repeatedly without hitting Excel again.
"""

from __future__ import annotations

import hashlib
import logging
import os
import tempfile
import uuid
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Set, Tuple

import networkx as nx
from openpyxl import load_workbook
from xlcalculator import ModelCompiler

from app.core.config import settings
from app.db.arango import (
    cells,
    companies,
    dependencies,
    log_audit,
    scorecard_models,
    scorecard_snapshots,
)
from app.engine.addressing import normalize_address, normalize_inputs
from app.engine.graph_evaluator import ModelArtifact, evaluate_model as run_evaluation

logger = logging.getLogger(__name__)


class CompilationError(Exception):
    """Raised when workbook compilation fails."""


class DAGValidationError(Exception):
    """Raised when dependency graph has cycles."""


def _compute_file_hash(path: str) -> str:
    """Return a stable SHA256 hash of the file contents."""
    hasher = hashlib.sha256()
    with open(path, "rb") as f:
        while True:
            chunk = f.read(8192)
            if not chunk:
                break
            hasher.update(chunk)
    return hasher.hexdigest()


def _validate_workbook(path: str) -> List[str]:
    """Validate required sheets exist in the workbook.

    Returns a list of missing required sheet names (empty means ok).
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    required = settings.MODEL_REQUIRED_SHEETS
    if not required:
        return []

    missing = [s for s in required if s not in wb.sheetnames]
    return missing


def _clean_workbook(path: str) -> str:
    """Clean an Excel workbook by removing problematic features.

    Removes:
      - Conditional formatting
      - Data validation
      - Named ranges
      - External links

    Returns path to cleaned temporary workbook.
    """
    logger.info(f"Cleaning workbook: {path}")

    try:
        wb = load_workbook(path)

        cleaned_count = 0
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            if hasattr(ws, "conditional_formatting") and ws.conditional_formatting:
                ws.conditional_formatting._cf_rules.clear()
                cleaned_count += 1
                logger.debug(f"Removed conditional formatting from {sheet_name}")

            if hasattr(ws, "data_validations") and ws.data_validations.dataValidation:
                ws.data_validations.dataValidation.clear()
                cleaned_count += 1
                logger.debug(f"Removed data validation from {sheet_name}")

        if wb.named_ranges:
            for _ in list(wb.named_ranges):
                wb.named_ranges.pop()
            cleaned_count += 1
            logger.debug("Removed all named ranges")

        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.close()
        wb.save(tmp.name)

        if cleaned_count > 0:
            logger.info(f"Cleaned workbook: removed {cleaned_count} problematic feature sets")

        return tmp.name

    except Exception as e:
        logger.warning(f"Failed to clean workbook: {e}. Proceeding with original.")
        return path


def _compile_xlsx(path: str) -> Tuple[Dict[str, Any], Dict[str, Any], Dict[str, Any]]:
    """Compile an Excel workbook into cells, formulas, and dependency graph."""

    cleaned_path = _clean_workbook(path)
    cleanup_temp = cleaned_path != path

    try:
        try:
            logger.info(f"Compiling Excel file: {cleaned_path}")
            compiler = ModelCompiler()
            archive = compiler.read_excel_file(cleaned_path)
            compiler.parse_archive(archive)
            compiler.model.build_code()
        except MemoryError as e:
            logger.error("MemoryError while compiling: %s", e)
            raise CompilationError(
                "The Excel file is too complex. Try removing unused sheets, "
                "simplifying formulas, or reducing the number of cells with formulas."
            ) from e
        except Exception as e:
            logger.exception("Failed to compile Excel file")
            raise CompilationError(f"Failed to compile Excel file: {e}") from e

        model = compiler.model

        if len(model.cells) > settings.MAX_CELLS_PER_MODEL:
            raise CompilationError(
                f"Model exceeds maximum cells ({settings.MAX_CELLS_PER_MODEL}). "
                f"Found {len(model.cells)} cells."
            )

        default_sheet = None
        if model.cells:
            first_addr = next(iter(model.cells.keys()))
            if "!" in first_addr:
                default_sheet = first_addr.split("!", 1)[0]
        if not default_sheet:
            raise CompilationError("Unable to determine default sheet name from workbook")

        compiled_cells: Dict[str, Any] = {}
        formulas: Dict[str, str] = {}

        for address, cell in model.cells.items():
            addr = normalize_address(address, default_sheet=default_sheet)

            formula = None
            depends_on: List[str] = []

            if getattr(cell, "formula", None) is not None:
                formula = getattr(cell.formula, "formula", None)
                raw_deps = getattr(cell.formula, "associated_cells", [])
                depends_on = sorted(
                    normalize_address(dep, default_sheet=default_sheet)
                    for dep in raw_deps
                )

            compiled_cells[addr] = {
                "address": addr,
                "sheet": addr.split("!", 1)[0],
                "row": getattr(cell, "row", None),
                "column": getattr(cell, "column", None),
                "formula": formula,
                "depends_on": depends_on,
                "value": getattr(cell, "value", None),
            }

            if formula:
                formulas[addr] = formula

        nodes = list(compiled_cells.keys())
        edges: List[Dict[str, str]] = []
        for addr, cell in compiled_cells.items():
            for dep in cell.get("depends_on", []):
                edges.append({"from": dep, "to": addr})

        node_set = set(nodes)
        for e in edges:
            if e["from"] not in node_set or e["to"] not in node_set:
                raise CompilationError(
                    f"Inconsistent dependency graph edge: {e['from']} -> {e['to']}"
                )

        G = nx.DiGraph()
        G.add_nodes_from(nodes)
        for e in edges:
            G.add_edge(e["from"], e["to"])

        if not nx.is_directed_acyclic_graph(G):
            cycles = list(nx.simple_cycles(G))
            raise CompilationError(f"Dependency graph contains a cycle: {cycles}")

        if len(edges) > settings.MAX_EDGES_PER_MODEL:
            raise CompilationError(
                f"Dependency graph exceeds maximum edges ({settings.MAX_EDGES_PER_MODEL}). "
                f"Found {len(edges)} edges."
            )

        graph = {"nodes": nodes, "edges": edges, "default_sheet": default_sheet}
        return compiled_cells, formulas, graph

    finally:
        if cleanup_temp and cleaned_path != path:
            try:
                os.remove(cleaned_path)
            except Exception:
                pass


def _validate_dag(graph: Dict[str, Any]) -> None:
    """Validate the dependency graph is acyclic (DAG)."""
    G = nx.DiGraph()
    for node in graph.get("nodes", []):
        G.add_node(node)
    for edge in graph.get("edges", []):
        G.add_edge(edge["from"], edge["to"])

    if not nx.is_directed_acyclic_graph(G):
        cycles = list(nx.simple_cycles(G))
        raise DAGValidationError(f"Dependency graph contains cycles: {cycles}")


def _identify_inputs_and_outputs(compiled_cells: Dict[str, Any]) -> Tuple[List[str], List[str]]:
    """Identify input and output cells.

    Inputs: cells without formulas (can be overridden)
    Outputs: cells with formulas (top-level results)
    """
    inputs: List[str] = []
    outputs: List[str] = []

    for addr, info in compiled_cells.items():
        if not info.get("formula"):
            inputs.append(addr)
        else:
            outputs.append(addr)

    return inputs, outputs


def _store_cells_and_dependencies(
    version_id: str,
    compiled_cells: Dict[str, Any],
    graph: Dict[str, Any],
) -> None:
    """Persist cells and graph edges in ArangoDB."""
    for addr, info in compiled_cells.items():
        key = f"{version_id}_{addr}"
        doc = {
            "_key": key,
            "model_id": version_id,
            "address": addr,
            "sheet": info.get("sheet"),
            "formula": info.get("formula"),
            "value": info.get("value"),
            "depends_on": info.get("depends_on", []),
        }
        if cells.has(key):
            cells.update(doc)
        else:
            cells.insert(doc)

    for edge in graph.get("edges", []):
        src = edge["from"]
        dst = edge["to"]
        edge_key = f"{version_id}_{src}->{dst}"
        edge_doc = {
            "_key": edge_key,
            "model_id": version_id,
            "_from": f"cells/{version_id}_{src}",
            "_to": f"cells/{version_id}_{dst}",
        }
        if not dependencies.has(edge_key):
            dependencies.insert(edge_doc)


def create_model_version(
    name: str,
    filepath: str,
    uploaded_by: str = "system",
    metadata: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """Register a new model version and return its metadata.

    This function does NOT compile the model; compilation is intended to run as a
    background task (see `compile_model_version`).
    """
    missing = _validate_workbook(filepath)
    if missing:
        raise ValueError(f"Missing required sheets: {missing}")

    file_hash = _compute_file_hash(filepath)
    version_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()

    version_label = datetime.now(timezone.utc).strftime("v%Y%m%d%H%M%S")

    # Immutable model document (each upload is a distinct version)
    doc = {
        "_key": version_id,
        "version_id": version_label,
        "name": name,
        "created_at": now,
        "uploaded_by": uploaded_by,
        "file_name": os.path.basename(filepath),
        "file_hash": file_hash,
        "status": "processing",
        "metadata": metadata or {},
        # model_artifact will be populated after successful compilation
        "model_artifact": None,
        "cell_count": 0,
        "formula_count": 0,
        "input_count": 0,
        "output_count": 0,
        "error": None,
    }

    scorecard_models.insert(doc)
    log_audit(
        "model.upload.request",
        user=uploaded_by,
        details={"version_id": version_id, "name": name},
    )

    return doc


def compile_model_version(version_id: str, filepath: str) -> Dict[str, Any]:
    """Compile an uploaded Excel model and persist the compiled artifact."""
    doc = scorecard_models.get(version_id)
    if not doc:
        raise ValueError(f"Model version not found: {version_id}")

    # If already active, avoid re-compiling
    if doc.get("status") == "active":
        return doc

    try:
        compiled_cells, formulas, graph = _compile_xlsx(filepath)
        _validate_dag(graph)

        inputs, outputs = _identify_inputs_and_outputs(compiled_cells)

        # Build artifact and graph metadata
        model_artifact = {
            "cells": compiled_cells,
            "formulas": formulas,
            "inputs": inputs,
            "outputs": outputs,
            "graph": {
                "nodes": graph.get("nodes", []),
                "edges": graph.get("edges", []),
                "default_sheet": graph.get("default_sheet"),
            },
            "graph_meta": {
                "node_count": len(graph.get("nodes", [])),
                "edge_count": len(graph.get("edges", [])),
            },
        }

        # Persist graph vertices and edges into ArangoDB
        _store_cells_and_dependencies(version_id, compiled_cells, graph)

        updates = {
            "_key": version_id,
            "model_artifact": model_artifact,
            "cell_count": len(compiled_cells),
            "formula_count": len(formulas),
            "input_count": len(inputs),
            "output_count": len(outputs),
            "status": "active",
            "compiled_at": datetime.now(timezone.utc).isoformat(),
            "error": None,
        }

        scorecard_models.update(updates)
        log_audit(
            "model.upload.complete",
            user=doc.get("uploaded_by"),
            details={"version_id": version_id},
        )
        return updates

    except Exception as e:
        logger.exception("Failed to compile model %s", version_id)
        scorecard_models.update({
            "_key": version_id,
            "status": "failed",
            "error": str(e),
        })
        log_audit(
            "model.upload.failed",
            user=doc.get("uploaded_by"),
            details={"version_id": version_id, "error": str(e)},
        )
        raise


def _get_model_doc(version_id: str) -> Dict[str, Any]:
    doc = scorecard_models.get(version_id)
    if not doc:
        raise ValueError(f"Model version not found: {version_id}")
    return doc


def get_model_version(version_id: str) -> Dict[str, Any]:
    """Load model metadata by version ID."""
    doc = _get_model_doc(version_id)
    return doc


def get_model_artifact(version_id: str) -> ModelArtifact:
    """Load compiled model artifact by version ID."""
    doc = _get_model_doc(version_id)
    artifact = doc.get("model_artifact")
    if not artifact:
        raise ValueError("Model artifact not available; compilation may still be pending.")
    return ModelArtifact(artifact)


def evaluate_model(
    version_id: str,
    overrides: Optional[Dict[str, Any]] = None,
) -> Tuple[Dict[str, Any], Dict[str, Any]]:
    """Evaluate a model using stored artifact (NO recompilation)."""
    doc = _get_model_doc(version_id)
    status = doc.get("status")
    if status != "active":
        raise ValueError(f"Model version not active (status={status})")

    # Load compiled artifact from persisted model (no Excel, no recompile)
    artifact = get_model_artifact(version_id)

    # Evaluate using the graph evaluator (performs topological, partial recompute)
    results, stats = run_evaluation(artifact, overrides)

    # Persist snapshot for audit / historical analysis
    snapshot = {
        "_key": str(uuid.uuid4()),
        "model_version_id": version_id,
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "overrides": overrides or {},
        "results": results,
        "stats": stats,
    }
    scorecard_snapshots.insert(snapshot)

    log_audit(
        "model.evaluate",
        details={"version_id": version_id, "overrides": len(overrides or {}), "stats": stats},
    )

    return results, stats


def list_model_versions() -> List[Dict[str, Any]]:
    """List all model versions stored in ArangoDB."""
    return scorecard_models.all()


def summarize_model(version_id: str) -> Dict[str, Any]:
    """Get a summary of a model version."""
    doc = _get_model_doc(version_id)
    artifact = doc.get("model_artifact") or {}
    cells = artifact.get("cells", {})
    formulas = artifact.get("formulas", {})
    inputs = artifact.get("inputs", [])
    outputs = artifact.get("outputs", [])
    graph_meta = artifact.get("graph_meta", {})

    return {
        "version_id": version_id,
        "name": doc.get("name"),
        "uploaded_at": doc.get("created_at"),
        "status": doc.get("status"),
        "file_hash": doc.get("file_hash"),
        "cell_count": len(cells),
        "formula_count": len(formulas),
        "input_count": len(inputs),
        "output_count": len(outputs),
        "graph": {
            "nodes": int(graph_meta.get("node_count", 0)),
            "edges": int(graph_meta.get("edge_count", 0)),
        },
    }


def get_graph(version_id: str) -> Dict[str, Any]:
    """Get dependency graph for visualization."""
    artifact = get_model_artifact(version_id)
    return artifact.graph_data


def set_active_model(version_id: str, company_id: str = "global") -> Dict[str, Any]:
    """Mark a model version as active for a company or globally."""

    _get_model_doc(version_id)  # Validate version exists

    now = datetime.now(timezone.utc).isoformat()
    doc = {
        "_key": company_id,
        "active_model_id": version_id,
        "updated_at": now,
    }

    if companies.has(company_id):
        companies.update(doc)
    else:
        companies.insert(doc)

    log_audit(
        "model.set_active",
        details={"version_id": version_id, "company_id": company_id},
    )

    return {"company_id": company_id, "active_model_id": version_id}


def get_active_model(company_id: str = "global") -> Optional[str]:
    """Return the active model id for the given company or global."""
    doc = companies.get(company_id)
    if not doc:
        return None
    return doc.get("active_model_id")
